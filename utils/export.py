"""Excel export: build a multi-sheet Strategy Report in memory."""
import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Palette ─────────────────────────────────────────────────────────────────
_NAVY    = "0D1F3C"
_ACCENT  = "4A90D9"
_LIGHT   = "EEF2F7"
_WHITE   = "FFFFFF"
_BORDER  = "CBD5E0"
_GREEN   = "D4EDDA"
_RED     = "FAD7D7"

_HEADER_FONT   = Font(name="Calibri", bold=True, color=_WHITE, size=11)
_TITLE_FONT    = Font(name="Calibri", bold=True, color=_NAVY,  size=13)
_BODY_FONT     = Font(name="Calibri", color="1A202C", size=10)
_HEADER_FILL   = PatternFill("solid", fgColor=_NAVY)
_ACCENT_FILL   = PatternFill("solid", fgColor=_ACCENT)
_ALT_FILL      = PatternFill("solid", fgColor=_LIGHT)
_CENTER        = Alignment(horizontal="center", vertical="center")
_LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_thin          = Side(style="thin", color=_BORDER)
_BORDER_STYLE  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(ws, row, col, text, fill=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font  = _HEADER_FONT
    cell.fill  = fill or _HEADER_FILL
    cell.alignment = _CENTER
    cell.border    = _BORDER_STYLE
    return cell


def _cell(ws, row, col, value, bold=False, alt=False, num_fmt=None, align=_CENTER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Calibri", bold=bold, color="1A202C", size=10)
    cell.alignment = align
    cell.border    = _BORDER_STYLE
    if alt:
        cell.fill = _ALT_FILL
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _autofit(ws, extra=4):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + extra, 50)


# ── Sheet builders ───────────────────────────────────────────────────────────

def _sheet_summary(wb, df, sel_brands, t_start, t_end, adj_c, adj_q, adj_s,
                   base_clicks, base_cr, base_aov, event_log):
    ws = wb.create_sheet("📊 Strategy Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = "TECH STRATEGY LAB — STRATEGY REPORT"
    c.font      = Font(name="Calibri", bold=True, color=_NAVY, size=14)
    c.alignment = _CENTER
    c.fill      = PatternFill("solid", fgColor=_LIGHT)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    c2 = ws["A2"]
    c2.value     = f"Generated: {date.today().strftime('%d %B %Y')}  |  Brands: {', '.join(b.title() for b in sorted(sel_brands))}"
    c2.font      = Font(name="Calibri", italic=True, color="6B7280", size=10)
    c2.alignment = _CENTER

    # Section: Trial window
    r = 4
    ws.merge_cells(f"A{r}:D{r}")
    c = ws.cell(r, 1, "TRIAL WINDOW")
    c.font = _TITLE_FONT; c.fill = _HEADER_FILL; c.font = _HEADER_FONT
    c.alignment = _CENTER

    pairs = [
        ("Trial Start", str(t_start)),
        ("Trial End",   str(t_end)),
        ("Observed Clicks (raw)",  f"{adj_c:,.0f}"),
        ("Observed Qty (raw)",     f"{adj_q:,.0f}"),
        ("Observed Sales (raw)",   f"€{adj_s:,.2f}"),
    ]
    for i, (k, v) in enumerate(pairs, start=r+1):
        ws.cell(i, 1, k).font = Font(name="Calibri", bold=True, color="4A5568", size=10)
        ws.cell(i, 1).alignment = _LEFT
        ws.cell(i, 2, v).alignment = _LEFT
        ws.cell(i, 2).font = _BODY_FONT

    # Section: Calibrated base
    r2 = r + len(pairs) + 2
    ws.merge_cells(f"A{r2}:D{r2}")
    c = ws.cell(r2, 1, "CALIBRATED BASE PARAMETERS")
    c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _CENTER

    base_pairs = [
        ("Daily Base Clicks",   f"{base_clicks:,.1f}"),
        ("Base Conversion Rate", f"{base_cr:.3%}"),
        ("Base AOV",            f"€{base_aov:,.2f}"),
    ]
    for i, (k, v) in enumerate(base_pairs, start=r2+1):
        ws.cell(i, 1, k).font = Font(name="Calibri", bold=True, color="4A5568", size=10)
        ws.cell(i, 1).alignment = _LEFT
        ws.cell(i, 2, v).alignment = _LEFT
        ws.cell(i, 2).font = _BODY_FONT

    # Section: Full-year projections
    r3 = r2 + len(base_pairs) + 2
    ws.merge_cells(f"A{r3}:D{r3}")
    c = ws.cell(r3, 1, "FULL-YEAR PROJECTION TOTALS")
    c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _CENTER

    has_sim = "Sales_Sim" in df.columns
    proj_pairs = [
        ("Total Clicks (Base)",  f"{df['Clicks_Base'].sum():,.0f}"),
        ("Total Qty (Base)",     f"{df['Qty_Base'].sum():,.0f}"),
        ("Total Sales (Base)",   f"€{df['Sales_Base'].sum():,.2f}"),
    ]
    if has_sim:
        proj_pairs += [
            ("Total Clicks (Simulation)",  f"{df['Clicks_Sim'].sum():,.0f}"),
            ("Total Qty (Simulation)",     f"{df['Qty_Sim'].sum():,.0f}"),
            ("Total Sales (Simulation)",   f"€{df['Sales_Sim'].sum():,.2f}"),
            ("Sales Uplift",               f"€{df['Sales_Sim'].sum() - df['Sales_Base'].sum():,.2f}"),
        ]
    for i, (k, v) in enumerate(proj_pairs, start=r3+1):
        ws.cell(i, 1, k).font = Font(name="Calibri", bold=True, color="4A5568", size=10)
        ws.cell(i, 1).alignment = _LEFT
        ws.cell(i, 2, v).alignment = _LEFT
        ws.cell(i, 2).font = _BODY_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10


def _sheet_projection(wb, df):
    ws = wb.create_sheet("📈 Monthly Projection")
    ws.sheet_view.showGridLines = False

    monthly = (
        df.groupby("Month")
          .agg(
              Clicks_Base=("Clicks_Base", "sum"),
              Qty_Base=("Qty_Base", "sum"),
              Sales_Base=("Sales_Base", "sum"),
              Clicks_Sim=("Clicks_Sim", "sum") if "Clicks_Sim" in df.columns else ("Clicks_Base", "sum"),
              Qty_Sim=("Qty_Sim", "sum")     if "Qty_Sim"    in df.columns else ("Qty_Base",    "sum"),
              Sales_Sim=("Sales_Sim", "sum") if "Sales_Sim"  in df.columns else ("Sales_Base",  "sum"),
          )
          .reset_index()
    )
    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    headers = ["Month", "Clicks Base", "Clicks Sim", "Clicks Δ%",
               "Qty Base", "Qty Sim", "Qty Δ%",
               "Sales Base (€)", "Sales Sim (€)", "Sales Δ%"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    ws.row_dimensions[1].height = 22

    for ri, row in monthly.iterrows():
        r = ri + 2
        alt = (ri % 2 == 1)
        m  = int(row["Month"])
        mn = month_names[m-1] if 1 <= m <= 12 else str(m)
        cb, cs = row["Clicks_Base"], row["Clicks_Sim"]
        qb, qs = row["Qty_Base"], row["Qty_Sim"]
        sb, ss = row["Sales_Base"], row["Sales_Sim"]
        dc = (cs - cb) / cb * 100 if cb > 0 else 0
        dq = (qs - qb) / qb * 100 if qb > 0 else 0
        ds = (ss - sb) / sb * 100 if sb > 0 else 0

        vals = [mn, cb, cs, dc, qb, qs, dq, sb, ss, ds]
        fmts = [None, "#,##0", "#,##0", "+0.0%;-0.0%",
                "#,##0", "#,##0", "+0.0%;-0.0%",
                "#,##0.00", "#,##0.00", "+0.0%;-0.0%"]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = _cell(ws, r, ci, v, alt=alt, num_fmt=fmt)
            if isinstance(v, float) and "Δ" in headers[ci-1]:
                if v > 0:
                    cell.fill = PatternFill("solid", fgColor="D4EDDA")
                elif v < 0:
                    cell.fill = PatternFill("solid", fgColor="FAD7D7")

    _autofit(ws)


def _sheet_events(wb, event_log):
    ws = wb.create_sheet("⚡ Events & Attribution")
    ws.sheet_view.showGridLines = False

    headers = ["#", "Type", "Scope", "Level", "Target / Period", "Lift / Details", "Metric"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    if not event_log:
        ws.cell(2, 1, "No events logged in this session.").font = _BODY_FONT
        return

    type_fills = {
        "shock":           PatternFill("solid", fgColor="FEF3C7"),
        "reapplied_shock": PatternFill("solid", fgColor="FEF3C7"),
        "custom_drag":     PatternFill("solid", fgColor="EDE9FE"),
        "swap":            PatternFill("solid", fgColor="DBEAFE"),
    }

    for ri, ev in enumerate(event_log):
        r   = ri + 2
        alt = (ri % 2 == 1)
        ev_type  = ev.get("type", "")
        scope    = ev.get("scope", "post_trial")
        level    = ev.get("level", "Monthly")
        target   = str(ev.get("target", ev.get("a", "")))
        lift     = ev.get("lift", ev.get("shape", ""))
        metric   = ev.get("metric", "All")
        fill     = type_fills.get(ev_type, PatternFill("solid", fgColor=_LIGHT) if alt else None)

        row_vals = [ri+1, ev_type, scope, level, target, str(lift), metric]
        for ci, v in enumerate(row_vals, 1):
            cell = _cell(ws, r, ci, v, align=_LEFT)
            if fill:
                cell.fill = fill

    _autofit(ws)


def _sheet_dna(wb, df):
    ws = wb.create_sheet("🧬 DNA Profile")
    ws.sheet_view.showGridLines = False

    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    monthly_dna = (
        df.groupby("Month")
          .agg(
              idx_clicks_pure=("idx_clicks_pure", "mean"),
              idx_cr_pure=("idx_cr_pure", "mean"),
              idx_aov_pure=("idx_aov_pure", "mean"),
              idx_clicks_work=("idx_clicks_work", "mean"),
              idx_cr_work=("idx_cr_work", "mean"),
              idx_aov_work=("idx_aov_work", "mean"),
          )
          .reset_index()
    )

    headers = ["Month",
               "Clicks DNA (Base)", "CR DNA (Base)", "AOV DNA (Base)",
               "Clicks DNA (Sim)",  "CR DNA (Sim)",  "AOV DNA (Sim)"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 1, ci, h)

    for ri, row in monthly_dna.iterrows():
        r   = ri + 2
        alt = (ri % 2 == 1)
        m   = int(row["Month"])
        mn  = month_names[m-1] if 1 <= m <= 12 else str(m)
        vals = [mn,
                row["idx_clicks_pure"], row["idx_cr_pure"], row["idx_aov_pure"],
                row["idx_clicks_work"], row["idx_cr_work"], row["idx_aov_work"]]
        for ci, v in enumerate(vals, 1):
            _cell(ws, r, ci, round(v, 4) if isinstance(v, float) else v,
                  alt=alt, num_fmt="0.0000" if ci > 1 else None)

    _autofit(ws)


# ── Public API ───────────────────────────────────────────────────────────────

def build_excel_report(df, event_log, sel_brands,
                        t_start, t_end, adj_c, adj_q, adj_s,
                        base_clicks, base_cr, base_aov) -> bytes:
    """Return an Excel workbook as bytes suitable for st.download_button."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    _sheet_summary(wb, df, sel_brands, t_start, t_end, adj_c, adj_q, adj_s,
                   base_clicks, base_cr, base_aov, event_log)
    _sheet_projection(wb, df)
    _sheet_events(wb, event_log)
    _sheet_dna(wb, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
